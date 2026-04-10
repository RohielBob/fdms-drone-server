import os
import sqlite3
import threading
from datetime import datetime
from typing import Optional

from fastapi import FastAPI
from fastapi.responses import HTMLResponse, FileResponse
from pydantic import BaseModel
from openpyxl import Workbook, load_workbook

from fastapi.middleware.cors import CORSMiddleware

import firebase_admin
from firebase_admin import credentials, db

# =====================================================
# CONFIG
# =====================================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

FIREBASE_CREDENTIALS = os.path.join(BASE_DIR, "firebase_key.json")
EXCEL_FILE = os.path.join(BASE_DIR, "drone_data.xlsx")
SQLITE_DB = os.path.join(BASE_DIR, "drone_data.db")

app = FastAPI(
    title="FDMS Drone API",
    description="API de surveillance et d’analyse des données de vol du drone du Groupe 6",
    version="1.0.0"
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

excel_lock = threading.Lock()

# =====================================================
# FIREBASE
# =====================================================
FIREBASE_DB_URL = "https://drone-fdm-project-groupe-6-default-rtdb.firebaseio.com/"
firebase_enabled = False

if os.path.exists(FIREBASE_CREDENTIALS):
    try:
        if not firebase_admin._apps:
            cred = credentials.Certificate(FIREBASE_CREDENTIALS)
            firebase_admin.initialize_app(cred, {
                "databaseURL": FIREBASE_DB_URL
            })
        firebase_enabled = True
    except Exception as e:
        print("Erreur Firebase :", e)
        firebase_enabled = False
else:
    print("firebase_key.json introuvable")

# =====================================================
# MODELE (FIX 422 ICI)
# =====================================================
class DroneData(BaseModel):
    Flight_ID: Optional[str] = "MISSION"
    Date: Optional[str] = None
    timestamp: int

    altitude: float
    vitesse: float

    ax: float
    ay: float
    az: float

    roll: float
    pitch: float
    yaw: float

    pression: float
    temperature: float

    batterie: Optional[float] = 100.0


# =====================================================
# EXCEL INIT
# =====================================================
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.title = "Données Drone"
    ws.append([
        "Flight_ID","Date","timestamp","altitude","vitesse",
        "ax","ay","az","roll","pitch","yaw",
        "pression","temperature","batterie"
    ])
    wb.save(EXCEL_FILE)

# =====================================================
# SQLITE INIT (FIX IMPORTANT)
# =====================================================
def init_db():
    conn = sqlite3.connect(SQLITE_DB)
    cursor = conn.cursor()

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS drone_data (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            Flight_ID TEXT,
            Date TEXT,
            timestamp INTEGER,
            altitude REAL,
            vitesse REAL,
            ax REAL,
            ay REAL,
            az REAL,
            roll REAL,
            pitch REAL,
            yaw REAL,
            pression REAL,
            temperature REAL,
            batterie REAL
        )
    """)

    conn.commit()
    conn.close()

init_db()


def get_db_connection():
    conn = sqlite3.connect(SQLITE_DB)
    conn.row_factory = sqlite3.Row
    return conn


# =====================================================
# ALERTES
# =====================================================
def generate_alerts(data_dict):
    alerts = []

    if data_dict.get("temperature", 0) < -50 or data_dict.get("temperature", 0) > 150:
        alerts.append("Température inhabituelle")

    if data_dict.get("altitude", 0) < -100 or data_dict.get("altitude", 0) > 10000:
        alerts.append("Altitude inhabituelle")

    if data_dict.get("vitesse", 0) < 0 or data_dict.get("vitesse", 0) > 300:
        alerts.append("Vitesse inhabituelle")

    if data_dict.get("pression", 0) < 300 or data_dict.get("pression", 0) > 1200:
        alerts.append("Pression inhabituelle")

    if data_dict.get("roll", 0) < -180 or data_dict.get("roll", 0) > 180:
        alerts.append("Roll inhabituel")

    if data_dict.get("pitch", 0) < -180 or data_dict.get("pitch", 0) > 180:
        alerts.append("Pitch inhabituel")

    if data_dict.get("yaw", 0) < -360 or data_dict.get("yaw", 0) > 360:
        alerts.append("Yaw inhabituel")

    return alerts


# =====================================================
# HOME
# =====================================================
@app.get("/", response_class=HTMLResponse)
def home():
    return """
<!DOCTYPE html>
<html lang="fr">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Accueil - FDMS Drone Intelligence</title>
    <script src="https://cdn.tailwindcss.com"></script>
    <link href="https://fonts.googleapis.com/css2?family=Poppins:wght@300;400;600;800&display=swap" rel="stylesheet">
    <link href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.0.0/css/all.min.css" rel="stylesheet">
    <style>
        body { font-family: 'Poppins', sans-serif; background-color: #f8fafc; color: #1e293b; }
        .glass-card { background: rgba(255, 255, 255, 0.8); backdrop-filter: blur(12px); border-radius: 30px; border: 1px solid rgba(255, 255, 255, 0.5); box-shadow: 0 10px 40px rgba(0,0,0,0.03); }
        .gradient-text { background: linear-gradient(90deg, #3b82f6, #60a5fa); -webkit-background-clip: text; -webkit-text-fill-color: transparent; }
        .mission-icon { width: 60px; height: 60px; background: #eff6ff; color: #3b82f6; display: flex; align-items: center; justify-content: center; border-radius: 18px; font-size: 24px; margin-bottom: 20px; }
        header { background: rgba(255, 255, 255, 0.95); backdrop-filter: blur(10px); position: fixed; width: 100%; top: 0; z-index: 1000; border-bottom: 1px solid #e2e8f0; }
        
        /* Style des boutons de la barre latérale */
        .side-btn {
            display: flex;
            align-items: center;
            gap: 12px;
            padding: 12px 18px;
            background: rgba(255, 255, 255, 0.05);
            border: 1px solid rgba(255, 255, 255, 0.1);
            border-radius: 12px;
            color: #cbd5e1;
            text-decoration: none;
            font-weight: 500;
            transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
            width: 100%;
            text-align: left;
        }

        .side-btn i { width: 20px; text-align: center; }

        .side-btn:hover {
            background: #3b82f6;
            color: white;
            transform: translateX(8px);
            box-shadow: 0 4px 15px rgba(59, 130, 246, 0.4);
        }

        .sidebar-open { left: 0 !important; }
    </style>
</head>
<body class="pt-24">

    <button onclick="toggleSidebar()" class="fixed top-5 left-5 z-[2000] bg-blue-600 text-white p-3 rounded-xl shadow-lg hover:bg-blue-700 transition-all">
        <i class="fas fa-bars"></i>
    </button>

    <div id="sidebar" class="fixed top-0 left-[-300px] w-[280px] h-full bg-slate-900 z-[1500] shadow-2xl transition-all duration-300 ease-in-out p-6 pt-20">
        <div class="flex flex-col gap-3">
            <a href="/dashboard" class="side-btn"><i class="fas fa-chart-line"></i> Dashboard</a>
            <a href="/docs" class="side-btn"><i class="fas fa-file-code"></i> API Docs</a>
            <a href="/latest-data" class="side-btn"><i class="fas fa-clock"></i> Latest Data</a>
            <a href="/graph-data" class="side-btn"><i class="fas fa-database"></i> All Data</a>
            <a href="/docs#/default/receive_drone_data_drone_data_post" class="side-btn"><i class="fas fa-satellite-dish"></i> Réception</a>
            <a href="https://console.firebase.google.com/" target="_blank" class="side-btn"><i class="fab fa-google"></i> Firebase</a>
            <a href="/export-excel" class="side-btn"><i class="fas fa-file-excel"></i> Export Excel</a>
            <a href="/health" class="side-btn"><i class="fas fa-heartbeat"></i> Health Check</a>
            <button onclick="resetData()" class="side-btn text-red-400 border-red-900/30 hover:bg-red-900/20"><i class="fas fa-trash"></i> Reset Data</button>
        </div>
    </div>

    <div id="overlay" onclick="toggleSidebar()" class="fixed inset-0 bg-black/50 hidden z-[1400]"></div>

    <header class="py-4 px-8 flex justify-between items-center">
        <div class="text-2xl font-bold text-slate-900 ml-12">FDMS<span class="text-blue-500"> G-06</span></div>
        <nav class="hidden md:block">
            <ul class="flex gap-8 list-none">
                <li><a href="/" class="text-blue-600 font-bold">Accueil</a></li>
            </ul>
        </nav>
        <a href="/dashboard" class="bg-blue-600 text-white px-6 py-2 rounded-full font-semibold hover:bg-blue-700 transition shadow-lg shadow-blue-200">Live Monitor</a>
    </header>

    <main class="max-w-6xl mx-auto px-6 mt-16 mb-20">
        <div class="text-center mb-16">
            <h1 class="text-6xl font-extrabold text-slate-900 mb-6 tracking-tight">FDMS <span class="gradient-text">Control Center</span></h1>
            <p class="text-slate-500 text-xl max-w-2xl mx-auto font-light">
                Bienvenue sur le serveur FDMS du <strong>Groupe 6</strong>. Cette station de contrôle avancée permet la surveillance télémétrique et l'analyse de données en temps réel d'un drone.
            </p>
        </div>

        <div class="grid grid-cols-1 md:grid-cols-3 gap-8 mb-16">
            <div class="glass-card p-8">
                <div class="mission-icon"><i class="fas fa-bolt"></i></div>
                <h3 class="text-xl font-bold mb-3">Temps Réel</h3>
                <p class="text-slate-500 text-sm">Visualisation instantanée des données : altitude, vitesse, température, batterie, pression, orientation et accélération.</p>
            </div>
            <div class="glass-card p-8">
                <div class="mission-icon"><i class="fas fa-database"></i></div>
                <h3 class="text-xl font-bold mb-3">Archivage</h3>
                <p class="text-slate-500 text-sm">Stockage sécurisé sur SQLite et Firebase, avec exportation automatique vers Excel pour analyse.</p>
            </div>
            <div class="glass-card p-8">
                <div class="mission-icon"><i class="fas fa-shield-alt"></i></div>
                <h3 class="text-xl font-bold mb-3">Sécurité</h3>
                <p class="text-slate-500 text-sm">Algorithmes de détection d'anomalies et alertes automatiques en cas de dépassement de seuils.</p>
            </div>
        </div>

        <div class="glass-card p-10 bg-slate-900 text-white border-none relative overflow-hidden">
            <div class="relative z-10 flex flex-col md:flex-row justify-between items-center gap-8">
                <div>
                    <h2 class="text-3xl font-bold mb-4">Prêt pour le décollage ?</h2>
                    <p class="text-slate-400">Accédez au panneau de contrôle pour voir les données en direct.</p>
                </div>
                <a href="/dashboard" class="bg-white text-slate-900 px-8 py-4 rounded-2xl font-bold hover:scale-105 transition shadow-xl">Accéder au Dashboard →</a>
            </div>
            <div class="absolute top-0 right-0 opacity-10 transform translate-x-1/4 -translate-y-1/4">
                <i class="fas fa-plane-departure text-[200px]"></i>
            </div>
        </div>
    </main>

    <footer class="text-center py-12 text-slate-400 text-sm border-t border-slate-200">
        <p>&copy; 2026 - Projet FDMS Ingénierie. Développé par le Groupe 6.</p>
    </footer>
    
    <script>
        function toggleSidebar() {
            const sidebar = document.getElementById('sidebar');
            const overlay = document.getElementById('overlay');
            sidebar.classList.toggle('sidebar-open');
            overlay.classList.toggle('hidden');
        }

        async function resetData() {
            if(confirm("Êtes-vous sûr de vouloir supprimer TOUTES les données ?")) {
                try {
                    const response = await fetch('/reset-data', { method: 'DELETE' });
                    if(response.ok) {
                        alert("Données réinitialisées avec succès !");
                        location.reload();
                    }
                } catch (error) {
                    alert("Erreur lors de la réinitialisation.");
                }
            }
        }
    </script>
</body>
</html>
"""

# =====================================================
# POST DRONE DATA (MISSION PLANNER SAFE)
# =====================================================
@app.post("/drone-data")
def receive_drone_data(data: DroneData):

    data_dict = data.dict()

    # sécurité batterie (évite crash)
    if data_dict.get("batterie") is None:
        data_dict["batterie"] = 100.0

    alerts = generate_alerts(data_dict)

    # =========================
    # SQLITE
    # =========================
    conn = get_db_connection()
    cursor = conn.cursor()

    # === DANS TON CODE SERVEUR ===
    # === VERSION CORRIGÉE DU INSERT ===
    cursor.execute("""
        INSERT INTO drone_data (
            Flight_ID, Date, timestamp, altitude, vitesse,
            ax, ay, az, roll, pitch, yaw,
            pression, temperature, batterie
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        data.Flight_ID, data.Date, data.timestamp, data.altitude, data.vitesse,
        data.ax, data.ay, data.az, data.roll, data.pitch, data.yaw,
        data.pression, data.temperature, data.batterie
    ))

    conn.commit()
    conn.close()

    # =========================
    # EXCEL
    # =========================
    try:
        with excel_lock:
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active

            ws.append([
                data.Flight_ID,
                data.Date,
                data.timestamp,
                data.altitude,
                data.vitesse,
                data.ax,
                data.ay,
                data.az,
                data.roll,
                data.pitch,
                data.yaw,
                data.pression,
                data.temperature,
                data.batterie
            ])

            wb.save(EXCEL_FILE)
    except:
        pass

    # =========================
    # FIREBASE
    # =========================
    if firebase_enabled:
        try:
            db.reference("drone_data").push(data_dict)
        except:
            pass

    return {
        "status": "ok",
        "alerts": alerts,
        "data": data_dict
    }


# =====================================================
# LATEST DATA (SAFE)
# =====================================================
@app.get("/latest-data")
def latest_data():
    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM drone_data ORDER BY id DESC LIMIT 1")
    row = cursor.fetchone()
    conn.close()

    if not row:
        return {"message": "Aucune donnée"}

    data = dict(row)
    data["alerts"] = generate_alerts(data)
    return data


# =====================================================
# GRAPH DATA
# =====================================================
@app.get("/graph-data")
def graph_data():
    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM drone_data ORDER BY id ASC")
    rows = cursor.fetchall()
    conn.close()

    return [dict(row) for row in rows]


# =====================================================
# EXPORT EXCEL
# =====================================================
@app.get("/export-excel")
def export_excel():
    return FileResponse(EXCEL_FILE)


# =====================================================
# RESET DATA
# =====================================================
@app.delete("/reset-data")
def reset_data():
    conn = sqlite3.connect(SQLITE_DB)
    cursor = conn.cursor()
    cursor.execute("DELETE FROM drone_data")
    conn.commit()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.append([
        "Flight_ID","Date","timestamp","altitude","vitesse",
        "ax","ay","az","roll","pitch","yaw",
        "pression","temperature","batterie"
    ])
    wb.save(EXCEL_FILE)

    return {"status": "ok"}

@app.get("/dashboard", response_class=HTMLResponse)
def dashboard():
    html_content = """
<!DOCTYPE html>
<html lang="fr">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>FDMS Dashboard Pro - Groupe 6</title>
    <script src="https://cdn.tailwindcss.com"></script>
    <script src="https://cdn.jsdelivr.net/npm/apexcharts"></script>
    <link href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.0.0/css/all.min.css" rel="stylesheet">
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700;900&display=swap" rel="stylesheet">

    <style>
        body { background: #f8fafc; font-family: 'Inter', sans-serif; }
        .glass-card { 
            background: white; 
            border-radius: 16px; 
            border: 1px solid #e2e8f0;
            box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.05);
        }
        ::-webkit-scrollbar { width: 6px; }
        ::-webkit-scrollbar-thumb { background: #cbd5e1; border-radius: 10px; }
    </style>
</head>
<body class="flex h-screen p-6 gap-6">

    <aside class="w-72 glass-card p-6 flex flex-col shrink-0">
        <div class="flex items-center gap-3 mb-10 px-2">
            <div class="bg-blue-600 w-10 h-10 rounded-xl flex items-center justify-center text-white">
                <i class="fas fa-plane"></i>
            </div>
            <h1 class="text-xl font-black text-slate-800">FDMS <span class="text-blue-600">PRO</span></h1>
        </div>

        <nav class="flex-1 space-y-2">
            <a href="/" class="flex items-center gap-4 text-slate-500 hover:bg-slate-50 p-4 rounded-xl transition-all font-semibold">
                <i class="fas fa-home w-5"></i> Accueil
            </a>
            <div class="flex items-center gap-4 bg-blue-600 text-white p-4 rounded-xl font-bold shadow-lg shadow-blue-200">
                <i class="fas fa-chart-pie w-5"></i> Dashboard
            </div>
            <a href="/latest-data" target="_blank" class="flex items-center gap-4 text-slate-500 hover:bg-slate-50 p-4 rounded-xl transition-all font-semibold">
                <i class="fas fa-bolt w-5 text-amber-500"></i> Flux JSON
            </a>
            <div class="pt-4 mt-4 border-t border-slate-100 space-y-2">
                <a href="/download_csv" class="flex items-center gap-4 text-emerald-600 hover:bg-emerald-50 p-4 rounded-xl transition-all font-bold border border-emerald-100">
                    <i class="fas fa-file-csv"></i> Export Données
                </a>
                
                <button onclick="confirmDelete()" class="w-full flex items-center gap-4 text-red-600 hover:bg-red-50 p-4 rounded-xl transition-all font-bold border border-red-100">
                    <i class="fas fa-trash-alt"></i> Supprimer Data
                </button>
            </div>
        </nav>
    </aside>

    <main class="flex-1 flex flex-col gap-6 overflow-y-auto">
        
        <div class="grid grid-cols-4 gap-4">
            <div class="glass-card p-4">
                <p class="text-xs font-bold text-slate-400 uppercase">Altitude</p>
                <div class="text-2xl font-black text-slate-800"><span id="card-alt">--</span> <small class="text-slate-400 text-sm">m</small></div>
            </div>
            <div class="glass-card p-4">
                <p class="text-xs font-bold text-slate-400 uppercase">Vitesse</p>
                <div class="text-2xl font-black text-slate-800"><span id="card-vit">--</span> <small class="text-slate-400 text-sm">m/s</small></div>
            </div>
            <div class="glass-card p-4">
                <p class="text-xs font-bold text-slate-400 uppercase">Batterie</p>
                <div class="text-2xl font-black text-slate-800"><span id="card-batt">--</span> <small class="text-slate-400 text-sm">%</small></div>
            </div>
            <div class="glass-card p-4">
                <p class="text-xs font-bold text-slate-400 uppercase">Température</p>
                <div class="text-2xl font-black text-slate-800"><span id="card-temp">--</span> <small class="text-slate-400 text-sm">°C</small></div>
            </div>
        </div>

        <div id="charts-container" class="space-y-6">
            </div>
    </main>

    <script>
        const commonOptions = (colors, title) => ({
            chart: { 
                type: 'line', 
                height: 300, 
                toolbar: { show: false }, 
                animations: { enabled: false },
                background: '#fff'
            },
            colors: Array.isArray(colors) ? colors : [colors],
            stroke: { width: 3, curve: 'smooth' },
            grid: { 
                borderColor: '#f1f5f9', 
                xaxis: { lines: { show: true } }, 
                yaxis: { lines: { show: true } } 
            },
            xaxis: { 
                type: 'datetime',
                labels: { 
                    datetimeUTC: false,
                    format: 'HH:mm:ss',
                    style: { colors: '#64748b', fontSize: '10px' } 
                },
                title: { text: 'Timestamp (Date & Heure)', style: { color: '#94a3b8' } },
                axisBorder: { show: false }
            },
            yaxis: { 
                labels: { style: { colors: '#64748b' } },
                title: { text: title, style: { color: '#94a3b8' } }
            },
            tooltip: { x: { format: 'dd MMM yyyy HH:mm:ss' } }
        });

        const chartConfigs = [
            { id: 'altitude', color: '#3b82f6', label: 'Altitude (m)', multi: false },
            { id: 'vitesse', color: '#f43f5e', label: 'Vitesse (m/s)', multi: false },
            { id: 'pression', color: '#6366f1', label: 'Pression (hPa)', multi: false },
            { id: 'temperature', color: '#f59e0b', label: 'Température (°C)', multi: false },
            { id: 'batterie', color: '#10b981', label: 'Batterie (%)', multi: false },
            { id: 'accel', color: ['#3b82f6', '#f43f5e', '#10b981'], label: 'Accélération (AX, AY, AZ)', multi: true, keys: ['ax', 'ay', 'az'] },
            { id: 'attitude', color: ['#8b5cf6', '#ec4899'], label: 'Attitude (Roll, Pitch)', multi: true, keys: ['roll', 'pitch'] },
            { id: 'yaw', color: '#475569', label: 'Yaw (Cap °)', multi: false }
        ];

        const charts = {};
        const container = document.getElementById('charts-container');

        chartConfigs.forEach(conf => {
            const div = document.createElement('div');
            div.className = "glass-card p-6";
            div.innerHTML = `<h3 class="text-sm font-bold text-slate-700 mb-4 uppercase flex items-center gap-2">
                <span class="w-1 h-4 rounded" style="background:${Array.isArray(conf.color) ? conf.color[0] : conf.color}"></span> ${conf.label}
            </h3><div id="chart-${conf.id}"></div>`;
            container.appendChild(div);
            charts[conf.id] = new ApexCharts(document.querySelector(`#chart-${conf.id}`), commonOptions(conf.color, conf.label));
            charts[conf.id].render();
        });

        async function refresh() {
            try {
                const response = await fetch('/graph-data');
                const data = await response.json();
                if (!data.length) return;

                const last = data[data.length - 1];
                document.getElementById('card-alt').innerText = last.altitude;
                document.getElementById('card-vit').innerText = last.vitesse;
                document.getElementById('card-batt').innerText = last.batterie;
                document.getElementById('card-temp').innerText = last.temperature;

                const mapData = (key) => data.map(d => ({ x: new Date(d.timestamp).getTime(), y: d[key] }));

                chartConfigs.forEach(conf => {
                    if (conf.multi) {
                        const series = conf.keys.map(k => ({ name: k.toUpperCase(), data: mapData(k) }));
                        charts[conf.id].updateSeries(series, false);
                    } else {
                        charts[conf.id].updateSeries([{ name: conf.id, data: mapData(conf.id) }], false);
                    }
                });
            } catch (e) { console.error(e); }
        }

        async function confirmDelete() {
            if (confirm("⚠️ Êtes-vous sûr de vouloir supprimer TOUTES les données ?")) {
                try {
                    const res = await fetch('/delete-all', { method: 'DELETE' });
                    if (res.ok) {
                        alert("Données supprimées.");
                        window.location.reload();
                    }
                } catch (e) { alert("Erreur."); }
            }
        }

        setInterval(refresh, 800);
    </script>
</body>
</html>
    """
    return HTMLResponse(content=html_content)
